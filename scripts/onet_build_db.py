# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl"]
# ///
"""Build a SQLite database from O*NET Excel files for fast indexed queries."""

from __future__ import annotations

import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

REFERENCES_DIR = Path(__file__).resolve().parent.parent / "references"
DB_PATH = REFERENCES_DIR / "onet.db"
VERSION_FILE = REFERENCES_DIR / ".version"


def _sanitize_table_name(filename: str) -> str:
    name = filename.replace(".xlsx", "")
    name = re.sub(r"[^a-zA-Z0-9]+", "_", name)
    name = name.strip("_").lower()
    return name


def _sanitize_column_name(col: str) -> str:
    col = re.sub(r"[^a-zA-Z0-9]+", "_", col)
    col = col.strip("_").lower()
    return col


def _guess_type(values: list) -> str:
    for v in values:
        if v is None:
            continue
        if isinstance(v, (int,)):
            return "INTEGER"
        if isinstance(v, (float,)):
            return "REAL"
        return "TEXT"
    return "TEXT"


def build_database(references_dir: Path | None = None, db_path: Path | None = None) -> Path:
    ref = references_dir or REFERENCES_DIR
    db = db_path or DB_PATH

    xlsx_files = sorted(ref.glob("*.xlsx"))
    if not xlsx_files:
        print(f"Error: No .xlsx files found in {ref}", file=sys.stderr)
        sys.exit(1)

    tmp_db = db.with_suffix(".tmp")
    if tmp_db.exists():
        tmp_db.unlink()

    conn = sqlite3.connect(str(tmp_db))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")

    total = len(xlsx_files)
    for i, xlsx_path in enumerate(xlsx_files, 1):
        table_name = _sanitize_table_name(xlsx_path.name)
        pct = i * 100 // total
        print(f"  [{pct:3d}%] {xlsx_path.name} -> {table_name}")

        wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            continue

        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if raw_headers is None:
            wb.close()
            continue

        headers = [str(h) for h in raw_headers]
        col_names = [_sanitize_column_name(h) for h in headers]

        # Deduplicate column names
        seen: dict[str, int] = {}
        deduped: list[str] = []
        for c in col_names:
            if c in seen:
                seen[c] += 1
                deduped.append(f"{c}_{seen[c]}")
            else:
                seen[c] = 0
                deduped.append(c)
        col_names = deduped

        # Read all rows and determine types
        all_rows = list(rows_iter)
        wb.close()

        col_types: list[str] = []
        for ci in range(len(headers)):
            sample = [r[ci] for r in all_rows[:100] if ci < len(r)]
            col_types.append(_guess_type(sample))

        col_defs = ", ".join(
            f'"{cn}" {ct}' for cn, ct in zip(col_names, col_types)
        )
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.execute(f'CREATE TABLE "{table_name}" ({col_defs})')

        if all_rows:
            placeholders = ", ".join("?" * len(col_names))
            insert_sql = f'INSERT INTO "{table_name}" VALUES ({placeholders})'

            def _normalize_row(row: tuple, ncols: int) -> tuple:
                r = list(row[:ncols])
                while len(r) < ncols:
                    r.append(None)
                return tuple(
                    v.isoformat() if hasattr(v, "isoformat") else v for v in r
                )

            batch = [_normalize_row(r, len(col_names)) for r in all_rows]
            conn.executemany(insert_sql, batch)

        conn.commit()

    # Build indices on common lookup columns
    _create_indices(conn)

    # FTS5 full-text search index on occupations for fast keyword search
    _create_fts_index(conn)

    # Store metadata
    version = ""
    if VERSION_FILE.exists():
        version = VERSION_FILE.read_text().strip()
    conn.execute(
        "CREATE TABLE IF NOT EXISTS _metadata (key TEXT PRIMARY KEY, value TEXT)"
    )
    conn.execute(
        "INSERT OR REPLACE INTO _metadata VALUES ('version', ?)", (version,)
    )
    conn.execute(
        "INSERT OR REPLACE INTO _metadata VALUES ('build_date', datetime('now'))"
    )

    # Store original-filename-to-table mapping
    conn.execute(
        "CREATE TABLE IF NOT EXISTS _file_map (filename TEXT PRIMARY KEY, table_name TEXT)"
    )
    for xlsx_path in xlsx_files:
        conn.execute(
            "INSERT OR REPLACE INTO _file_map VALUES (?, ?)",
            (xlsx_path.name, _sanitize_table_name(xlsx_path.name)),
        )

    conn.commit()
    conn.execute("VACUUM")
    conn.close()

    if db.exists():
        db.unlink()
    tmp_db.rename(db)

    return db


def _create_indices(conn: sqlite3.Connection) -> None:
    cursor = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE '_%'"
    )
    tables = [row[0] for row in cursor.fetchall()]

    for table in tables:
        col_cursor = conn.execute(f'PRAGMA table_info("{table}")')
        columns = [row[1] for row in col_cursor.fetchall()]

        if "o_net_soc_code" in columns:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "idx_{table}_code" '
                f'ON "{table}" ("o_net_soc_code")'
            )
        if "element_id" in columns:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "idx_{table}_element" '
                f'ON "{table}" ("element_id")'
            )
        if "scale_id" in columns:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "idx_{table}_scale" '
                f'ON "{table}" ("scale_id")'
            )
        if "task_id" in columns:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "idx_{table}_task" '
                f'ON "{table}" ("task_id")'
            )
        if "commodity_code" in columns:
            conn.execute(
                f'CREATE INDEX IF NOT EXISTS "idx_{table}_commodity" '
                f'ON "{table}" ("commodity_code")'
            )

    conn.commit()


def _create_fts_index(conn: sqlite3.Connection) -> None:
    conn.execute("DROP TABLE IF EXISTS occupation_fts")
    conn.execute("""
        CREATE VIRTUAL TABLE occupation_fts USING fts5(
            o_net_soc_code,
            title,
            description,
            content='occupation_data',
            content_rowid='rowid',
            tokenize='porter unicode61'
        )
    """)
    conn.execute("""
        INSERT INTO occupation_fts(rowid, o_net_soc_code, title, description)
        SELECT rowid, o_net_soc_code, title, description FROM occupation_data
    """)

    # Also index alternate titles for broader search coverage
    conn.execute("DROP TABLE IF EXISTS alternate_titles_fts")
    conn.execute("""
        CREATE VIRTUAL TABLE alternate_titles_fts USING fts5(
            o_net_soc_code,
            title,
            alternate_title,
            content='alternate_titles',
            content_rowid='rowid',
            tokenize='porter unicode61'
        )
    """)
    conn.execute("""
        INSERT INTO alternate_titles_fts(rowid, o_net_soc_code, title, alternate_title)
        SELECT rowid, o_net_soc_code, title, alternate_title FROM alternate_titles
    """)
    conn.commit()


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Build SQLite database from O*NET Excel files.",
    )
    parser.add_argument(
        "--data-dir",
        help="Path to the references/ directory (default: auto-detect)",
    )
    parser.add_argument(
        "--output",
        "-o",
        help="Output database path (default: references/onet.db)",
    )
    args = parser.parse_args()

    ref = Path(args.data_dir) if args.data_dir else REFERENCES_DIR
    out = Path(args.output) if args.output else DB_PATH

    if not ref.exists():
        print(f"Error: Directory not found: {ref}", file=sys.stderr)
        sys.exit(1)

    print(f"Building SQLite database from {ref}/*.xlsx ...")
    db = build_database(ref, out)
    size_mb = db.stat().st_size / (1024 * 1024)
    print(f"\nDone: {db} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    main()
